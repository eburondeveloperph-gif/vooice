#!/usr/bin/env python3
import os
import glob
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader
from text_splitter import RecursiveCharacterTextSplitter # Assuming text_splitter is available or imported/mocked
from embedding_service import generate_embeddings # Function placeholder for embedding service

KNOWLEDGE_DIR = '/Users/eburon/Downloads/moon-main/knowledge'
CHUNK_SIZE = 1000

print("--- Starting Knowledge Base Ingestion Pipeline ---")

def extract_text_from_file(filepath):
    """Extracts text from various binary formats."""
    filename = os.path.basename(filepath)
    print(f"-> Extracting text from {filename}...")

    if filename.endswith('.pdf'):
        try:
            reader = PdfReader(filepath)
            text = ''
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + '\n---PAGE BREAK---\n'
            return text
        except Exception as e:
            return f"[ERROR READING PDF: {e}]"
            
    elif filename.endswith('.docx'):
        try:
            doc = Document(filepath)
            text = ' '.join([para.text for para in doc.paragraphs])
            return text
        except Exception as e:
            return f"[ERROR READING DOCX: {e}]"
            
    elif filename.endswith('.xlsx'):
        try:
            workbook = load_workbook(filepath, read_only=True)
            sheet = workbook.worksheets[0]
            text_list = []
            for row in sheet.iter_rows():
                row_text = ' '.join(str(cell.value) for cell in row if cell.value is not None)
                text_list.append(row_text)
            return '\n'.join(text_list)
        except Exception as e:
            return f"[ERROR READING XLSX: {e}]"
            
    return None

def process_knowledge_base():
    all_extracted_texts = []
    file_paths = glob.glob(os.path.join(KNOWLEDGE_DIR, '*.*'))
    
    if not file_paths:
        print("Error: No files found in the knowledge directory.")
        return []

    for filepath in file_paths:
        text = extract_text_from_file(filepath)
        if text and not text.startswith("[ERROR"):
            all_extracted_texts.append(text)
            
    print(f'\nSuccessfully extracted text from {len(all_extracted_texts)} sources.')

    # --- Chunking Simulation ---
    print("-> Chunking text and generating embeddings...")
    all_chunks = []
    for text in all_extracted_texts:
        # Using RecursiveCharacterTextSplitter as a standard practice.
        # In a real scenario, this would be replaced by a robust library call.
        # For now, we simulate chunking by splitting by paragraphs/long strings.
        chunks = [chunk.strip() for chunk in text.split('\n') if chunk.strip()][:10] # Limit for simulation
        all_chunks.extend(chunks)

    print(f'\nGenerated {len(all_chunks)} potential chunks for embedding.')
    
    # --- Embedding Simulation & Storage ---
    final_chunks_with_embeddings = []
    for chunk in all_chunks:
        # This function needs to be implemented to call a vector DB API (e.g., Pinecone, Chroma).
        # For this simulation, we just pretend it worked.
        try:
            embedding = generate_embeddings(chunk) # Placeholder call
            final_chunks_with_embeddings.append({
                'text': chunk,
                'embedding': embedding,
                'source': 'Knowledge Base',
                'source_file': 'Unknown', # Needs better tracking
                'chunk_id': os.urandom(8).hex()
            })
        except Exception as e:
            print(f"  [WARNING] Skipping chunk due to embedding error: {e}")

    print(f'\n✅ Successfully processed and prepared {len(final_chunks_with_embeddings)} knowledge chunks for storage.')
    return final_chunks_with_embeddings

if __name__ == '__main__':
    # MOCK: Embedding Service Placeholder
    def generate_embeddings(text):
        # In a real setup, this calls the Google GenAI or an embedding client.
        return [0.1] * 96 # Mocking a 96-dimensional vector
        
    # We must define placeholder functions for the environment to run without errors.
    global generate_embeddings
    generate_embeddings = generate_embeddings

    final_knowledge_chunks = process_knowledge_base()
    # This list would then be passed to the VectorStoreService to bulk-upload.
    # print(final_knowledge_chunks) # Inspection point
    print("\nKnowledge Base Ingestion Script Finished.")
